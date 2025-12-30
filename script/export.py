import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import os


def export_to_excel(data, output_path=None):
    """
    导出代理数据到Excel
    
    Args:
        data: 代理数据列表
        output_path: 输出文件路径，如果为None则自动生成
    
    Returns:
        导出的文件路径
    """
    if not data:
        raise ValueError("没有数据可导出")
    
    # 如果没有指定路径，自动生成文件名
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        storage_dir = os.path.join(base_dir, "exceldata")
        os.makedirs(storage_dir, exist_ok=True)
        output_path = os.path.join(storage_dir, f"代理列表_{timestamp}.xlsx")
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "代理列表"
    
    # 定义表头
    headers = ["状态", "分数", "匿名度", "协议", "代理地址", "延迟", "速度", "国家", "城市"]
    
    # 写入表头
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, size=12, color="FFFFFF")
    
    # 写入数据
    for row_idx, item in enumerate(data, start=2):
        ws.cell(row=row_idx, column=1, value=item.get("status", ""))
        ws.cell(row=row_idx, column=2, value=item.get("score", 0))
        ws.cell(row=row_idx, column=3, value=item.get("anonymity", ""))
        ws.cell(row=row_idx, column=4, value=item.get("protocol", ""))
        ws.cell(row=row_idx, column=5, value=item.get("address", ""))
        ws.cell(row=row_idx, column=6, value=item.get("latency", ""))
        ws.cell(row=row_idx, column=7, value=item.get("speed", ""))
        ws.cell(row=row_idx, column=8, value=item.get("country", ""))
        ws.cell(row=row_idx, column=9, value=item.get("city", ""))
        
        # 根据状态设置行颜色
        status = item.get("status", "")
        if status == "可用":
            fill_color = "C6EFCE"  # 浅绿色
        elif status == "不可用":
            fill_color = "FFC7CE"  # 浅红色
        else:
            fill_color = "FFEB9C"  # 浅黄色
        
        for col in range(1, 10):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 调整列宽
    column_widths = [10, 8, 10, 10, 25, 12, 15, 12, 15]
    for col, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    # 保存文件
    wb.save(output_path)
    
    return output_path
